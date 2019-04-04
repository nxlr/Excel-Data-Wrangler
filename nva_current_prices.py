import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

import os.path
from os import path

# List of 22 districts of Haryana
district_list = ['Ambala','Bhiwani','Charkhi Dadri','Faridabad','Fatehabad',
'Gurugram','Hisar','Jhajjar','Jind','Kaithal','Karnal','Kurukshetra','Mahendragarh',
'Mewat','Palwal','Panchkula','Panipat','Rewari','Rohtak','Sirsa','Sonipat','Yamunanagar']

# New workbook filepath
wb_filepath="./Haryana DDP Workbook.xlsx"

# Check if workbook already exists, create one if it does not exist
if not path.exists(wb_filepath):

    # Create new workbook for district wise data 
    wb = Workbook()

    # Create new excel sheets for each district
    for district in district_list:
        # create district sheet if it does not already exist in the workbook
        if not district in wb.sheetnames:
            wb.create_sheet(district)
        # save workbook
        wb.save(wb_filepath)

    # Remove Intial Blank Sheet if it exists 
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
        wb.save(wb_filepath)

# filepath of original data file which contains yearwise data
filepath = "./mycopy_Haryana DDP Estimates Work March 27 2019.xlsx"

# original workbook is loaded in rawbook variable
rawbook = load_workbook(filepath, data_only=True)
# load the new workbook where data will be copied
wb = load_workbook(wb_filepath)

from copy import copy

# copy the S.No. and Item names in each district sheet
for district_sheet in wb.worksheets:
    # Heading of A1 and A2 in each sheet
    district_sheet['D77'].value = "Net Value Added ( at Current Prices )"
    # copy style
    fontStyle = Font(size = "15", bold = True)
    alignment = Alignment(horizontal='center', vertical='center')
    #fill = PatternFill("solid", fgColor="FFFFFF")

    district_sheet['D77'].font = fontStyle
    district_sheet["D77"].alignment = alignment
    #district_sheet["D1"].fill = fill

    district_sheet['D78'].font = fontStyle
    district_sheet['D78'].alignment = alignment
    #district_sheet['D2'].fill = fill

    # copy the S.No. and Item names in each district sheet
    for col in range(1,3):
        for row in range(80,114):
            district_sheet.cell(row=row, column=col).value = rawbook['2011-12'].cell(row=row, column=col).value
            # copy cell font style as well
            district_sheet.cell(row=row, column=col).font = copy(rawbook['2011-12'].cell(row=row, column=col).font)
    
    # adjust column B's width
    district_sheet.column_dimensions['B'].width = 40
    # save workbook
    wb.save(wb_filepath)


# col = 3 for Ambala in all yearwise sheets initially
col = 3

flag = False

# write year-wise data in district sheets
for district_sheet in wb.worksheets:
    # yearwise column for district sheet, intitially j = 3 means year 2011-12
    j = 3
    
    # Special Case
    if district_sheet.title == "Charkhi Dadri": 
        ws = rawbook["2017-18"]
        district_sheet.cell(row=79, column=3).value = "Charkhi Dadri"
        district_sheet.cell(row=80, column=3).value = "2017-18"
        for row in range(81,114):
            district_sheet.cell(row=row, column=j).value = ws.cell(row=row, column=5).value
            district_sheet.column_dimensions[get_column_letter(j)].width = 15
            # copy cell font and number format style
            district_sheet.cell(row=row, column=j).font = copy(ws.cell(row=row, column=col).font)
            district_sheet.cell(row=row, column=j).number_format = copy(ws.cell(row=row, column=col).number_format)
            # use flag to indicate that Charkhi Dadri has been covered
            flag = True
        continue
    
    # loop over rawbook sheets to get data for each district
    for year_sheet in rawbook.worksheets:
         # assign district name to corresponding column in district sheet
        district_sheet.cell(row=79, column=j).value = district_sheet.title
        # assign year name to corresponding column in district sheet
        district_sheet.cell(row=80, column=j).value = year_sheet.title
        # temporary variable to retain col value
        tmp = col
        # because one additional column is in the 2017-18 sheet
        if flag==True and year_sheet.title=="2017-18":
            col = col+1
        # get and assign data for that year column    
        for row in range(81,114):
            district_sheet.cell(row=row, column=j).value = rawbook[year_sheet.title].cell(row=row, column=col).value
            district_sheet.column_dimensions[get_column_letter(j)].width = 15
            # copy cell font and number format style
            district_sheet.cell(row=row, column=j).font = copy(rawbook[year_sheet.title].cell(row=row, column=col).font)
            district_sheet.cell(row=row, column=j).number_format = copy(rawbook[year_sheet.title].cell(row=row, column=col).number_format)
        j = j + 1
        col = tmp
    # increment col to get column data for next district 
    col = col + 1
    # save workbook
    wb.save(wb_filepath)




# Make the Code for Charkhi Dadri dynamic
# Write code for all sections of each sheet - GVA and NVA at both current and constant prices
