import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from copy import copy
import os.path
from os import path
import time

start = time.time()

def prCyan(skk): print("\033[96m {}\033[00m" .format(skk)) 
prCyan("\t\t Preparing Your File. Please Wait...")

# Create Intermediate File
# os.system("start.py")

# List of 22 districts of Haryana
district_list = ['Ambala','Bhiwani','Charkhi Dadri','Faridabad','Fatehabad',
'Gurugram','Hisar','Jhajjar','Jind','Kaithal','Karnal','Kurukshetra','Mahendragarh',
'Mewat','Palwal','Panchkula','Panipat','Rewari','Rohtak','Sirsa','Sonipat','Yamunanagar']

# New workbook filepath
out_filepath="./Final_Data_Horizontal.xlsx"
# filepath of Intermediate file which this code imports data from
filepath = "./Haryana DDP Workbook.xlsx"

var = 'fsheet'

# remove old output file
if path.exists(out_filepath):
    os.remove(out_filepath)

# Create new output file
out = Workbook()
# Create new worksheet at first position in output file
fsheet = out.create_sheet(var, 0)

# Remove Intial Blank Sheet if it exists 
if 'Sheet' in out.sheetnames:
    out.remove(out['Sheet']) 
# save workbook
out.save(out_filepath)

# original workbook file is loaded in databook variable
databook = load_workbook(filepath, data_only=True)
# load the new output file where data will be copied
out = load_workbook(out_filepath)

fsheet = out[var]

for case in range(1,5):
    if case == 1:
        # Headings
        fsheet['F1'].value = "Gross Value Added ( at Current Prices )"
        # define style for headings
        fontStyle = Font(size = "15", bold = True)
        alignment = Alignment(horizontal='center', vertical='center')
        # set style for headings
        fsheet['F1'].font = fontStyle
        fsheet["F1"].alignment = alignment
        fsheet['F2'].font = fontStyle
        fsheet['F2'].alignment = alignment

        # copy the S.No. and Item names 
        for col in range(1,3):
            for row in range(4,38):
                fsheet.cell(row=row, column=col).value = databook['Ambala'].cell(row=row, column=col).value
                # copy cell font style as well
                fsheet.cell(row=row, column=col).font = copy(databook['Ambala'].cell(row=row, column=col).font)
        
        # adjust column B's width
        fsheet.column_dimensions['B'].width = 40
        # save workbook
        out.save(out_filepath)

        # temporary variable to keep track of column to paste in
        tmpCol = 3  

        for district_sheet in databook.worksheets:

            if district_sheet.title == "Charkhi Dadri":
                tmpCol = 17
                colRange = 4
            else:
                colRange = 10

            for col in range(3, colRange):
                # convert column number to ASCII format
                fromCol = get_column_letter(col)
                toCol = get_column_letter(tmpCol)
                # set column width as original one
                fsheet.column_dimensions[toCol].width = district_sheet.column_dimensions[fromCol].width
                for row in range(3,38):
                    # copy value row-wise for each column
                    fsheet.cell(row=row, column=tmpCol).value = district_sheet.cell(row=row, column=col).value
                    # copy cell font and number format style
                    fsheet.cell(row=row, column=tmpCol).font = copy(district_sheet.cell(row=row, column=col).font)
                    fsheet.cell(row=row, column=tmpCol).number_format = copy(district_sheet.cell(row=row, column=col).number_format)                
                tmpCol += 1

            out.save(out_filepath)    

    if case == 2:
        # Headings
        fsheet['F39'].value = "Gross Value Added ( at Constant Prices )"
        # define style for headings
        fontStyle = Font(size = "15", bold = True)
        alignment = Alignment(horizontal='center', vertical='center')
        # set style for headings
        fsheet['F39'].font = fontStyle
        fsheet["F39"].alignment = alignment
        fsheet['F39'].font = fontStyle
        fsheet['F39'].alignment = alignment

        # copy the S.No. and Item names 
        for col in range(1,3):
            for row in range(42,76):
                fsheet.cell(row=row, column=col).value = databook['Ambala'].cell(row=row, column=col).value
                # copy cell font style as well
                fsheet.cell(row=row, column=col).font = copy(databook['Ambala'].cell(row=row, column=col).font)
        
        # save workbook
        out.save(out_filepath)

        # temporary variable to keep track of column to paste in
        tmpCol = 3  

        for district_sheet in databook.worksheets:

            if district_sheet.title == "Charkhi Dadri":
                tmpCol = 17
                colRange = 4
            else:
                colRange = 10

            for col in range(3, colRange):
                for row in range(41,76):
                    # copy value row-wise for each column
                    fsheet.cell(row=row, column=tmpCol).value = district_sheet.cell(row=row, column=col).value
                    # copy cell font and number format style
                    fsheet.cell(row=row, column=tmpCol).font = copy(district_sheet.cell(row=row, column=col).font)
                    fsheet.cell(row=row, column=tmpCol).number_format = copy(district_sheet.cell(row=row, column=col).number_format)                
                tmpCol += 1

            out.save(out_filepath)

    if case == 3:
        # Headings
        fsheet['F77'].value = "Net Value Added ( at Current Prices )"
        # define style for headings
        fontStyle = Font(size = "15", bold = True)
        alignment = Alignment(horizontal='center', vertical='center')
        # set style for headings
        fsheet['F77'].font = fontStyle
        fsheet["F77"].alignment = alignment
        fsheet['F77'].font = fontStyle
        fsheet['F77'].alignment = alignment

        # copy the S.No. and Item names 
        for col in range(1,3):
            for row in range(80,114):
                fsheet.cell(row=row, column=col).value = databook['Ambala'].cell(row=row, column=col).value
                # copy cell font style as well
                fsheet.cell(row=row, column=col).font = copy(databook['Ambala'].cell(row=row, column=col).font)
        
        # save workbook
        out.save(out_filepath)

        # temporary variable to keep track of column to paste in
        tmpCol = 3  

        for district_sheet in databook.worksheets:

            if district_sheet.title == "Charkhi Dadri":
                tmpCol = 17
                colRange = 4
            else:
                colRange = 10

            for col in range(3, colRange):
                for row in range(79,114):
                    # copy value row-wise for each column
                    fsheet.cell(row=row, column=tmpCol).value = district_sheet.cell(row=row, column=col).value
                    # copy cell font and number format style
                    fsheet.cell(row=row, column=tmpCol).font = copy(district_sheet.cell(row=row, column=col).font)
                    fsheet.cell(row=row, column=tmpCol).number_format = copy(district_sheet.cell(row=row, column=col).number_format)                
                tmpCol += 1

            out.save(out_filepath)

    if case == 4:
        # Headings
        fsheet['F115'].value = "Net Value Added ( at Constant Prices )"
        # define style for headings
        fontStyle = Font(size = "15", bold = True)
        alignment = Alignment(horizontal='center', vertical='center')
        # set style for headings
        fsheet['F115'].font = fontStyle
        fsheet["F115"].alignment = alignment
        fsheet['F115'].font = fontStyle
        fsheet['F115'].alignment = alignment

        # copy the S.No. and Item names 
        for col in range(1,3):
            for row in range(118, 152):
                fsheet.cell(row=row, column=col).value = databook['Ambala'].cell(row=row, column=col).value
                # copy cell font style as well
                fsheet.cell(row=row, column=col).font = copy(databook['Ambala'].cell(row=row, column=col).font)
        
        # save workbook
        out.save(out_filepath)

        # temporary variable to keep track of column to paste in
        tmpCol = 3  

        for district_sheet in databook.worksheets:

            if district_sheet.title == "Charkhi Dadri":
                tmpCol = 17
                colRange = 4
            else:
                colRange = 10

            for col in range(3, colRange):
                for row in range(117, 152):
                    # copy value row-wise for each column
                    fsheet.cell(row=row, column=tmpCol).value = district_sheet.cell(row=row, column=col).value
                    # copy cell font and number format style
                    fsheet.cell(row=row, column=tmpCol).font = copy(district_sheet.cell(row=row, column=col).font)
                    fsheet.cell(row=row, column=tmpCol).number_format = copy(district_sheet.cell(row=row, column=col).number_format)                
                tmpCol += 1

            out.save(out_filepath)
             
end = time.time()

exec_time = end - start
print("\n\t\t Execution Time : ", exec_time, "\n")